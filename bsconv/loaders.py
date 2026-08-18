"""Load any supported statement file into a uniform grid of cell values.

Supported inputs:
  * .xlsx / .xlsm            openpyxl
  * .xls (real BIFF)         xlrd
  * .xls that is really HTML  BeautifulSoup  (Xalq Bank's iBank export)
  * .htm / .html             BeautifulSoup
  * .csv / .tsv              csv module with delimiter and encoding sniffing

Everything downstream sees `Grid`: a rectangular list of rows of raw Python
values. Dates and numbers are left as native types where the source provides
them, because that is strictly more information than their string form.
"""

from __future__ import annotations

import csv
import io
import os
import re
import zipfile
from dataclasses import dataclass, field
from typing import Any, Sequence

Row = list[Any]


class UnsupportedFileError(ValueError):
    """Raised when a file cannot be read as a statement at all."""


@dataclass
class Grid:
    """A single worksheet, normalised to a rectangular list of rows."""

    rows: list[Row] = field(default_factory=list)
    sheet_name: str = ""
    source_format: str = ""

    @property
    def nrows(self) -> int:
        return len(self.rows)

    @property
    def ncols(self) -> int:
        return max((len(r) for r in self.rows), default=0)

    def cell(self, r: int, c: int) -> Any:
        if 0 <= r < len(self.rows) and 0 <= c < len(self.rows[r]):
            return self.rows[r][c]
        return None

    def row_text(self, r: int) -> str:
        """Row rendered as one string, with merge artefacts collapsed.

        A value spread across a merged range repeats once per column; emitting
        "Лицевой счет No 2020... " five times would break metadata regexes and
        name extraction, so consecutive duplicates are folded.
        """
        from .normalize import clean_text
        parts: list[str] = []
        for value in self.rows[r]:
            if value in (None, ""):
                continue
            text = clean_text(value)
            if text and (not parts or parts[-1] != text):
                parts.append(text)
        return " ".join(parts).strip()

    def is_blank(self, r: int) -> bool:
        return all(v is None or str(v).strip() == "" for v in self.rows[r])


# --------------------------------------------------------------------------
# format sniffing
# --------------------------------------------------------------------------

_HTML_HEAD_RE = re.compile(rb"^\s*(<\?xml|<!DOCTYPE|<html|<HTML|<table|<TABLE)", re.I)
_OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
_ZIP_MAGIC = b"PK\x03\x04"


def sniff_format(data: bytes, filename: str = "") -> str:
    """Return one of: xlsx, xls, html, csv."""
    head = data[:512]
    if head.startswith(_ZIP_MAGIC):
        return "xlsx"
    if head.startswith(_OLE2_MAGIC):
        return "xls"
    if _HTML_HEAD_RE.match(head) or b"<table" in head.lower():
        return "html"
    ext = os.path.splitext(filename)[1].lower()
    if ext in {".csv", ".tsv", ".txt"}:
        return "csv"
    if ext in {".htm", ".html"}:
        return "html"
    # Last resort: if it decodes as text with delimiters, treat as csv.
    for enc in ("utf-8-sig", "cp1251"):
        try:
            text = data[:4096].decode(enc)
        except UnicodeDecodeError:
            continue
        if text.count(";") + text.count(",") + text.count("\t") > 3:
            return "csv"
    raise UnsupportedFileError(
        f"Cannot determine file format for {filename or '<bytes>'}"
    )


# --------------------------------------------------------------------------
# per-format loaders
# --------------------------------------------------------------------------

def _col_letters_to_index(letters: str) -> int:
    """'A' -> 0, 'B' -> 1, ..., 'Z' -> 25, 'AA' -> 26, ..."""
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1


def _merged_ranges(data: bytes, worksheet_path: str | None) -> list[tuple[int, int, int, int]]:
    """Read <mergeCells> straight from the sheet XML.

    Read-only worksheets expose no merged-cell API at all (ws.merged_cells
    doesn't exist), so this bypasses openpyxl for just this one piece, using
    the archive path openpyxl itself resolved for the worksheet. Returns
    0-indexed (min_row, min_col, max_row, max_col) tuples; degrades to no
    merges found rather than raising if the path is missing or unreadable.
    """
    if not worksheet_path:
        return []
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as z:
            xml = z.read(worksheet_path).decode("utf-8", errors="replace")
    except (KeyError, zipfile.BadZipFile, OSError):
        return []
    ranges = []
    for m in re.finditer(r'<mergeCell ref="([A-Z]+)(\d+):([A-Z]+)(\d+)"', xml):
        c1, r1, c2, r2 = m.groups()
        ranges.append((int(r1) - 1, _col_letters_to_index(c1),
                       int(r2) - 1, _col_letters_to_index(c2)))
    return ranges


def _load_xlsx(data: bytes) -> list[Grid]:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    grids: list[Grid] = []
    for ws in wb.worksheets:
        # Some exports declare a bogus <dimension> (seen as low as "A1" on a
        # real 19k-row, 3MB export) that read-only mode trusts by default,
        # silently truncating the sheet to almost nothing. Force a real scan
        # of the XML instead of trusting that tag.
        ws.reset_dimensions()
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        width = max((len(r) for r in rows), default=0)
        rows = [r + [None] * (width - len(r)) for r in rows]

        # Merged cells carry their value only in the top-left cell. Spread it
        # HORIZONTALLY so a header label sitting in a merged range is visible
        # in the column its data actually occupies.
        #
        # Vertical spreading is deliberately NOT done: layouts that merge a
        # cell down several rows (Ipak Yo'li, Ipoteka) would turn one
        # transaction into two or three identical-looking data rows.
        for (r, c1, _r2, c2) in _merged_ranges(data, getattr(ws, "_worksheet_path", None)):
            if not (0 <= r < len(rows)):
                continue
            value = rows[r][c1] if c1 < len(rows[r]) else None
            if value is None:
                continue
            for c in range(c1, c2 + 1):
                if c < len(rows[r]) and rows[r][c] is None:
                    rows[r][c] = value
        grids.append(Grid(rows=rows, sheet_name=ws.title, source_format="xlsx"))
    wb.close()
    return grids


def _load_xls(data: bytes) -> list[Grid]:
    import xlrd

    book = xlrd.open_workbook(file_contents=data)
    grids: list[Grid] = []
    for sheet in book.sheets():
        rows: list[Row] = []
        for r in range(sheet.nrows):
            row: Row = []
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate.xldate_as_datetime(value, book.datemode)
                    except Exception:  # noqa: BLE001 - malformed serial
                        pass
                elif cell.ctype == xlrd.XL_CELL_EMPTY:
                    value = None
                row.append(value)
            rows.append(row)
        # Horizontal-only spread of merged regions, same reasoning as xlsx.
        for rlo, _rhi, clo, chi in getattr(sheet, "merged_cells", []):
            value = rows[rlo][clo] if rlo < len(rows) and clo < len(rows[rlo]) else None
            if value in (None, ""):
                continue
            for c in range(clo, min(chi, len(rows[rlo]))):
                if rows[rlo][c] in (None, ""):
                    rows[rlo][c] = value
        grids.append(Grid(rows=rows, sheet_name=sheet.name, source_format="xls"))
    return grids


def _decode_html(data: bytes) -> str:
    m = re.search(rb'charset=["\']?([\w\-]+)', data[:2048], re.I)
    encodings = []
    if m:
        encodings.append(m.group(1).decode("ascii", "ignore"))
    encodings += ["utf-8", "cp1251", "koi8-r", "latin-1"]
    for enc in encodings:
        try:
            return data.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
    return data.decode("latin-1", "replace")


def _load_html(data: bytes) -> list[Grid]:
    from bs4 import BeautifulSoup

    soup = BeautifulSoup(_decode_html(data), "html.parser")
    tables = soup.find_all("table")
    if not tables:
        raise UnsupportedFileError("HTML file contains no <table> element")

    def table_rows(table: Any) -> list[Row]:
        out: list[Row] = []
        for tr in table.find_all("tr", recursive=True):
            # Skip rows belonging to a nested table; they are emitted separately.
            if tr.find_parent("table") is not table:
                continue
            cells: Row = []
            for td in tr.find_all(["td", "th"], recursive=False):
                # <br> is structural in these exports: inside one cell it
                # separates the correspondent header, the counterparty name
                # and the payment purpose. Keep it as a newline so the field
                # splitter can recover all three.
                for br in td.find_all("br"):
                    br.replace_with("\n")
                text = td.get_text("", strip=False).replace("\xa0", " ")
                text = "\n".join(
                    re.sub(r"[ \t]+", " ", line).strip() for line in text.split("\n")
                ).strip()
                text = re.sub(r"\n{2,}", "\n", text)
                span = int(td.get("colspan") or 1)
                cells.append(text)
                cells.extend([None] * (span - 1))
            if cells:
                out.append(cells)
        return out

    # Concatenate all top-level tables into one grid: these exports split the
    # header block and the transaction block into sibling tables.
    rows: list[Row] = []
    for table in tables:
        if table.find_parent("table") is not None:
            continue
        rows.extend(table_rows(table))
    if not rows:
        for table in tables:
            rows.extend(table_rows(table))
    return [Grid(rows=rows, sheet_name="html", source_format="html")]


def _load_csv(data: bytes) -> list[Grid]:
    text = None
    for enc in ("utf-8-sig", "utf-8", "cp1251", "latin-1"):
        try:
            text = data.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise UnsupportedFileError("Cannot decode CSV file")
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=";,\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ";" if text.count(";") > text.count(",") else ","
    rows = [list(r) for r in csv.reader(io.StringIO(text), delimiter=delimiter)]
    return [Grid(rows=rows, sheet_name="csv", source_format="csv")]


_LOADERS = {
    "xlsx": _load_xlsx,
    "xls": _load_xls,
    "html": _load_html,
    "csv": _load_csv,
}


def load_grids(source: str | bytes | os.PathLike, filename: str = "") -> list[Grid]:
    """Load a file path or raw bytes into one Grid per worksheet."""
    if isinstance(source, (str, os.PathLike)):
        filename = filename or str(source)
        with open(source, "rb") as fh:
            data = fh.read()
    else:
        data = bytes(source)

    fmt = sniff_format(data, filename)
    try:
        grids = _LOADERS[fmt](data)
    except UnsupportedFileError:
        raise
    except Exception as exc:  # noqa: BLE001
        # A .xls that is really HTML, or an .xlsx that is really a .xls.
        for alt, loader in _LOADERS.items():
            if alt == fmt:
                continue
            try:
                grids = loader(data)
                break
            except Exception:  # noqa: BLE001
                continue
        else:
            raise UnsupportedFileError(
                f"Failed to read {filename or '<bytes>'} as {fmt}: {exc}"
            ) from exc

    return [g for g in grids if g.nrows]


def trim(grid: Grid) -> Grid:
    """Drop trailing all-blank rows and columns."""
    rows = grid.rows
    while rows and all(v is None or str(v).strip() == "" for v in rows[-1]):
        rows.pop()
    width = 0
    for row in rows:
        for i, value in enumerate(row):
            if value is not None and str(value).strip():
                width = max(width, i + 1)
    grid.rows = [list(r[:width]) + [None] * (width - len(r[:width])) for r in rows]
    return grid


def load(source: str | bytes | os.PathLike, filename: str = "") -> list[Grid]:
    return [trim(g) for g in load_grids(source, filename)]


__all__ = ["Grid", "load", "load_grids", "sniff_format", "UnsupportedFileError", "trim"]
